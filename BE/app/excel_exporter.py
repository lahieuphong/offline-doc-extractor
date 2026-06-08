import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


EXCEL_COLUMNS = [
    ("docId", "Mã tài liệu"),
    ("arcDocCode", "Mã hồ sơ tài liệu"),
    ("maintenance", "Chế độ bảo trì/lưu trữ"),
    ("typeName", "Loại văn bản"),
    ("codeNumber", "Số văn bản"),
    ("codeNotation", "Ký hiệu văn bản"),
    ("issuedDate", "Ngày ban hành"),
    ("organName", "Cơ quan ban hành"),
    ("subject", "Trích yếu nội dung"),
    ("language", "Ngôn ngữ"),
    ("numberOfPage", "Số trang"),
    ("inforSign", "Thông tin ký"),
    ("keyword", "Từ khóa"),
    ("mode", "Chế độ xử lý"),
    ("confidenceLevel", "Mức độ tin cậy"),
    ("autograph", "Chữ ký/khối ký"),
    ("format", "Định dạng tệp"),
    ("process", "Quy trình xử lý"),
    ("riskRecovery", "Rủi ro khôi phục"),
    ("riskRecoveryStatus", "Trạng thái rủi ro khôi phục"),
    ("description", "Mô tả"),
    ("isCan", "Cờ kiểm tra"),
]


def stringify_cell(value: Any) -> Optional[Union[str, int, float]]:
    if value is None:
        return None

    if isinstance(value, (str, int, float)):
        return value

    return json.dumps(value, ensure_ascii=False)


def export_results_to_excel(results: List[Dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ket qua boc tach"

    headers = ["STT"] + [label for _, label in EXCEL_COLUMNS]
    sheet.append(headers)

    header_font = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
    body_font = Font(name="Times New Roman", size=12)
    header_fill = PatternFill(fill_type="solid", fgColor="0B3A8E")
    thin_border = Border(
        left=Side(style="thin", color="D7DEEA"),
        right=Side(style="thin", color="D7DEEA"),
        top=Side(style="thin", color="D7DEEA"),
        bottom=Side(style="thin", color="D7DEEA"),
    )

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, item in enumerate(results, start=1):
        row = [row_index]

        for key, _label in EXCEL_COLUMNS:
            row.append(stringify_cell(item.get(key)))

        sheet.append(row)

    # Widths tuned for 22-field administrative metadata.  Long fields are kept
    # fully in-cell (wrapped) instead of being truncated.
    column_widths = {
        "A": 8, "B": 20, "C": 22, "D": 24, "E": 16, "F": 12, "G": 18, "H": 14,
        "I": 24, "J": 72, "K": 14, "L": 10, "M": 38, "N": 46, "O": 14, "P": 38,
        "Q": 34, "R": 12, "S": 46, "T": 24, "U": 28, "V": 58, "W": 24,
    }
    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.row_dimensions[1].height = 36
    for row in sheet.iter_rows():
        if row[0].row > 1:
            sheet.row_dimensions[row[0].row].height = 72
        for cell in row:
            if cell.row > 1:
                cell.font = body_font
                cell.border = thin_border
                if cell.column in {1, 6, 8, 11, 12, 18}:
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
